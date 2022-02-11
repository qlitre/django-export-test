from django.views import generic
from .models import Post
from django.http import HttpResponse
import csv
from .forms import PostSearchForm
import openpyxl
from django_pandas.io import read_frame


class Top(generic.ListView):
    """
    一覧ページ
    """
    model = Post

    def get_queryset(self):
        queryset = super().get_queryset()
        self.form = form = PostSearchForm(self.request.GET or None)

        if form.is_valid():
            key_word = form.cleaned_data.get('key_word')
            if key_word:
                for word in key_word.split():
                    queryset = queryset.filter(title__icontains=word)

            category = form.cleaned_data.get('category')
            if category:
                queryset = queryset.filter(category=category)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_form'] = self.form

        return context


def export(request):
    """
    csvエクスポート
    検索に応じて絞り込む場合
    """
    posts = Post.objects.all()
    # 検索に応じて絞り込む
    key_word = request.GET.get('key_word')
    if key_word:
        for word in key_word.split():
            posts = posts.filter(title__icontains=word)

    category = request.GET.get('category')
    if category:
        posts = posts.filter(category=category)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="posts.csv"'

    writer = csv.writer(response)
    header = ['pk', 'title', 'category']
    writer.writerow(header)

    for post in posts:
        writer.writerow([post.pk, post.title, post.category])
    return response


def export_excel(request):
    """Excel形式でのexport"""
    wb = openpyxl.Workbook()
    ws = wb.active
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=posts.xlsx'

    ws.cell(1, 1).value = 'pk'
    ws.cell(1, 2).value = 'title'
    ws.cell(1, 3).value = 'category'
    k = 2
    for post in Post.objects.all():
        ws.cell(k, 1).value = post.pk
        ws.cell(k, 2).value = post.title
        # .nameをつけないとバグ
        ws.cell(k, 3).value = post.category.name
        k += 1

    wb.save(response)

    return response


def export_pandas_dataframe(request):
    """
    pandas DataFrameをエクスポート
    """
    df = read_frame(qs=Post.objects.all(), fieldnames=['pk', 'title', 'category'])
    response = HttpResponse(content_type='text/csv; charset=utf8')
    response['Content-Disposition'] = 'attachment; filename=posts.csv'
    # 実際に使う場合はここで何らかの処理が入る
    # とりあえず列名だけ変更
    df = df.rename(columns={'pk': '番号', 'title': 'タイトル', 'category': 'カテゴリ'})

    df.to_csv(path_or_buf=response, encoding='utf-8', index=None)

    return response
